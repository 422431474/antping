#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DNS IPv6 è§£æçˆ¬è™«
ä» antping.com/dns ç½‘ç«™è·å–åŸŸåçš„ IPv6 (AAAAè®°å½•) è§£æç»“æœ

ä¼˜åŒ–ç­–ç•¥ï¼š
- ä¿æŒå•ä¸ªé¡µé¢ï¼Œç›´æ¥ä¿®æ”¹è¾“å…¥æ¡†å†…å®¹è¿›è¡ŒæŸ¥è¯¢
- ä¸é‡å¤åŠ è½½é¡µé¢ï¼Œæé«˜æ•ˆç‡
- æ·»åŠ è¯¦ç»†æ—¥å¿—è®°å½•
"""

import asyncio
import re
import time
import json
import ipaddress
import logging
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.utils import get_column_letter

# é…ç½®æ—¥å¿—
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(f'dns_crawler_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class DNSIPv6Crawler:
    def __init__(self, excel_path: str, headless: bool = True, use_proxy: bool = True, 
                 proxy_host: str = "127.0.0.1", proxy_port: int = 7890,
                 requests_per_ip: int = 10):
        self.excel_path = excel_path
        self.headless = headless
        self.base_url = "https://antping.com/dns"
        self.results = {}
        self.page_initialized = False  # æ ‡è®°é¡µé¢æ˜¯å¦å·²åˆå§‹åŒ–ï¼ˆé€‰æ‹©äº†AAAAè®°å½•ç±»å‹ï¼‰
        
        # ä»£ç†é…ç½®
        self.use_proxy = use_proxy
        self.proxy_host = proxy_host
        self.proxy_port = proxy_port
        self.proxy_url = f"http://{proxy_host}:{proxy_port}" if use_proxy else None
        
        # IPåˆ‡æ¢é…ç½®
        self.requests_per_ip = requests_per_ip  # æ¯ä¸ªIPå¤„ç†çš„è¯·æ±‚æ•°
        self.current_request_count = 0  # å½“å‰IPå·²å¤„ç†çš„è¯·æ±‚æ•°
        
        # æ–­ç‚¹ç»­ä¼ é…ç½®
        self.progress_file = excel_path.replace('.xlsx', '_progress.json')
        self.last_failed_index = None  # è®°å½•å¤±è´¥æ—¶çš„ç´¢å¼•
        
        # Clash APIé…ç½®ï¼ˆç”¨äºåˆ‡æ¢èŠ‚ç‚¹ï¼‰
        self.clash_api_url = "http://127.0.0.1:9090"
        self.clash_secret = ""  # å¦‚æœæœ‰å¯†ç ï¼Œå¡«å†™è¿™é‡Œ
        self.proxy_group = "ğŸ”° èŠ‚ç‚¹é€‰æ‹©"  # ä»£ç†ç»„åç§°
        
    def check_proxy_available(self) -> bool:
        """æ£€æŸ¥ä»£ç†æ˜¯å¦å¯ç”¨"""
        import socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex((self.proxy_host, self.proxy_port))
        sock.close()
        return result == 0
    
    def get_clash_proxies(self) -> list:
        """è·å–Clashå¯ç”¨çš„ä»£ç†èŠ‚ç‚¹åˆ—è¡¨"""
        import urllib.request
        try:
            url = f"{self.clash_api_url}/proxies/{urllib.parse.quote(self.proxy_group)}"
            headers = {}
            if self.clash_secret:
                headers['Authorization'] = f'Bearer {self.clash_secret}'
            
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=5) as response:
                data = json.loads(response.read().decode())
                # è¿‡æ»¤æ‰éèŠ‚ç‚¹çš„é€‰é¡¹
                all_proxies = data.get('all', [])
                valid_proxies = [p for p in all_proxies if not any(x in p for x in ['æµé‡', 'å¥—é¤', 'é‡ç½®', 'ç›´è¿', 'è‡ªåŠ¨', 'æ•…éšœ'])]
                return valid_proxies
        except Exception as e:
            logger.warning(f"è·å–Clashä»£ç†åˆ—è¡¨å¤±è´¥: {e}")
            return []
    
    def switch_clash_proxy(self, proxy_name: str) -> bool:
        """åˆ‡æ¢Clashä»£ç†èŠ‚ç‚¹"""
        import urllib.request
        import urllib.parse
        try:
            url = f"{self.clash_api_url}/proxies/{urllib.parse.quote(self.proxy_group)}"
            headers = {'Content-Type': 'application/json'}
            if self.clash_secret:
                headers['Authorization'] = f'Bearer {self.clash_secret}'
            
            data = json.dumps({'name': proxy_name}).encode('utf-8')
            req = urllib.request.Request(url, data=data, headers=headers, method='PUT')
            with urllib.request.urlopen(req, timeout=5) as response:
                logger.info(f"å·²åˆ‡æ¢åˆ°ä»£ç†èŠ‚ç‚¹: {proxy_name}")
                return True
        except Exception as e:
            logger.warning(f"åˆ‡æ¢Clashä»£ç†å¤±è´¥: {e}")
            return False
    
    def get_next_proxy(self) -> str:
        """è·å–ä¸‹ä¸€ä¸ªä»£ç†èŠ‚ç‚¹"""
        proxies = self.get_clash_proxies()
        if not proxies:
            return None
        
        # éšæœºé€‰æ‹©ä¸€ä¸ªèŠ‚ç‚¹
        import random
        return random.choice(proxies)
        
    async def init_browser(self):
        """åˆå§‹åŒ–æµè§ˆå™¨ï¼ˆæ— ç—•æ¨¡å¼+éšæœºæŒ‡çº¹ï¼‰"""
        logger.info("æ­£åœ¨å¯åŠ¨æµè§ˆå™¨ï¼ˆæ— ç—•æ¨¡å¼ï¼‰...")
        self.playwright = await async_playwright().start()
        
        # é…ç½®ä»£ç†
        launch_options = {"headless": self.headless}
        
        # æ— ç—•æ¨¡å¼ + éšæœºæŒ‡çº¹é…ç½®
        import random
        user_agents = [
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:121.0) Gecko/20100101 Firefox/121.0",
        ]
        
        viewports = [
            {"width": 1920, "height": 1080},
            {"width": 1440, "height": 900},
            {"width": 1536, "height": 864},
            {"width": 1366, "height": 768},
        ]
        
        context_options = {
            "user_agent": random.choice(user_agents),
            "viewport": random.choice(viewports),
            "locale": random.choice(["zh-CN", "zh-TW", "en-US"]),
            "timezone_id": random.choice(["Asia/Shanghai", "Asia/Hong_Kong", "Asia/Taipei"]),
        }
        
        if self.use_proxy:
            if self.check_proxy_available():
                context_options["proxy"] = {"server": self.proxy_url}
                logger.info(f"ä½¿ç”¨ä»£ç†: {self.proxy_url}")
            else:
                logger.warning(f"ä»£ç† {self.proxy_url} ä¸å¯ç”¨ï¼Œå°†ä¸ä½¿ç”¨ä»£ç†")
                self.use_proxy = False
        
        self.browser = await self.playwright.chromium.launch(**launch_options)
        # ä½¿ç”¨ new_context åˆ›å»ºæ— ç—•æµè§ˆä¸Šä¸‹æ–‡ï¼ˆPlaywrighté»˜è®¤å°±æ˜¯æ— ç—•çš„ï¼‰
        self.context = await self.browser.new_context(**context_options)
        self.page = await self.context.new_page()
        logger.info(f"æµè§ˆå™¨å¯åŠ¨æˆåŠŸï¼ŒUA: {context_options['user_agent'][:50]}...")
        
    async def restart_browser_for_new_ip(self):
        """é‡å¯æµè§ˆå™¨å¹¶åˆ‡æ¢ä»£ç†èŠ‚ç‚¹"""
        logger.info("=" * 40)
        logger.info(f"å·²å¤„ç† {self.requests_per_ip} ä¸ªè¯·æ±‚ï¼Œåˆ‡æ¢ä»£ç†èŠ‚ç‚¹...")
        logger.info("=" * 40)
        
        # å°è¯•é€šè¿‡Clash APIåˆ‡æ¢èŠ‚ç‚¹
        next_proxy = self.get_next_proxy()
        if next_proxy:
            self.switch_clash_proxy(next_proxy)
            await asyncio.sleep(2)  # ç­‰å¾…åˆ‡æ¢ç”Ÿæ•ˆ
        else:
            logger.warning("æ— æ³•è·å–ä»£ç†èŠ‚ç‚¹åˆ—è¡¨ï¼Œå°†ç»§ç»­ä½¿ç”¨å½“å‰èŠ‚ç‚¹")
        
        # å…³é—­å½“å‰æµè§ˆå™¨
        await self.context.close()
        await self.browser.close()
        
        # ç­‰å¾…ä¸€æ®µæ—¶é—´
        logger.info("ç­‰å¾…3ç§’...")
        await asyncio.sleep(3)
        
        # é‡æ–°åˆå§‹åŒ–æµè§ˆå™¨
        context_options = {}
        if self.use_proxy and self.check_proxy_available():
            context_options["proxy"] = {"server": self.proxy_url}
            
        self.browser = await self.playwright.chromium.launch(headless=self.headless)
        self.context = await self.browser.new_context(**context_options)
        self.page = await self.context.new_page()
        
        # é‡ç½®çŠ¶æ€
        self.page_initialized = False
        self.current_request_count = 0
        
        logger.info("æµè§ˆå™¨å·²é‡å¯ï¼Œç»§ç»­çˆ¬å–...")
        
    async def close_browser(self):
        """å…³é—­æµè§ˆå™¨"""
        logger.info("æ­£åœ¨å…³é—­æµè§ˆå™¨...")
        await self.context.close()
        await self.browser.close()
        await self.playwright.stop()
        logger.info("æµè§ˆå™¨å·²å…³é—­")
        
    def read_domains_from_excel(self) -> list:
        """ä»Excelè¯»å–åŸŸååˆ—è¡¨"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        domains = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            domain = row[0].value
            if domain:
                domains.append(domain.strip())
        wb.close()
        return domains
    
    def save_progress(self, current_index: int, results: dict):
        """ä¿å­˜å½“å‰è¿›åº¦åˆ°æ–‡ä»¶"""
        progress_data = {
            'last_index': current_index,
            'timestamp': datetime.now().isoformat(),
            'results': results
        }
        with open(self.progress_file, 'w', encoding='utf-8') as f:
            json.dump(progress_data, f, ensure_ascii=False, indent=2)
        logger.info(f"è¿›åº¦å·²ä¿å­˜åˆ°: {self.progress_file} (ç´¢å¼•: {current_index})")
    
    def load_progress(self) -> tuple:
        """åŠ è½½ä¹‹å‰çš„è¿›åº¦"""
        if Path(self.progress_file).exists():
            try:
                with open(self.progress_file, 'r', encoding='utf-8') as f:
                    progress_data = json.load(f)
                last_index = progress_data.get('last_index', 0)
                results = progress_data.get('results', {})
                timestamp = progress_data.get('timestamp', '')
                logger.info(f"æ‰¾åˆ°è¿›åº¦æ–‡ä»¶ï¼Œä¸Šæ¬¡åœæ­¢äºç´¢å¼• {last_index}ï¼Œæ—¶é—´: {timestamp}")
                return last_index, results
            except Exception as e:
                logger.warning(f"åŠ è½½è¿›åº¦æ–‡ä»¶å¤±è´¥: {e}")
        return 0, {}
    
    def is_valid_ipv6(self, addr: str) -> bool:
        """éªŒè¯æ˜¯å¦æ˜¯æœ‰æ•ˆçš„IPv6åœ°å€"""
        try:
            # è¿‡æ»¤æ‰å¤ªçŸ­çš„åœ°å€ï¼ˆçœŸå®IPv6åœ°å€è‡³å°‘æœ‰15ä¸ªå­—ç¬¦ï¼Œå¦‚ ::1 è¿™ç§æç«¯æƒ…å†µé™¤å¤–ï¼‰
            if len(addr) < 10:
                return False
            # æ’é™¤ä¸€äº›æ˜æ˜¾ä¸æ˜¯IPv6çš„æ¨¡å¼
            if addr.endswith('::') and len(addr) < 15:
                return False
            ipaddress.IPv6Address(addr)
            return True
        except:
            return False
    
    def extract_ipv6_addresses(self, text: str) -> set:
        """ä»æ–‡æœ¬ä¸­æå–æ‰€æœ‰æœ‰æ•ˆçš„IPv6åœ°å€"""
        # æ›´ç²¾ç¡®çš„IPv6åœ°å€æ­£åˆ™è¡¨è¾¾å¼
        # ä¸»è¦åŒ¹é…ç±»ä¼¼ 240e:6b0:ab0:11:1::1086 è¿™ç§æ ¼å¼
        # è¦æ±‚è‡³å°‘æœ‰3ä¸ªå†’å·åˆ†éš”çš„éƒ¨åˆ†
        ipv6_pattern = r'\b([0-9a-fA-F]{1,4}(?::[0-9a-fA-F]{0,4}){2,7})\b'
        
        matches = re.findall(ipv6_pattern, text)
        valid_ipv6 = set()
        
        for match in matches:
            # é¢å¤–éªŒè¯ï¼šå¿…é¡»åŒ…å«è‡³å°‘3ä¸ªå†’å·
            if match.count(':') >= 3 and self.is_valid_ipv6(match):
                valid_ipv6.add(match)
        
        return valid_ipv6
    
    async def wait_for_loading_complete(self, timeout: int = 10):
        """ç­‰å¾…é¡µé¢åŠ è½½åŠ¨ç”»æ¶ˆå¤±"""
        try:
            # ç­‰å¾…spinneræ¶ˆå¤±
            spinner = self.page.locator('.ant-spin-spinning')
            await spinner.wait_for(state="hidden", timeout=timeout * 1000)
        except:
            pass  # å¦‚æœæ²¡æœ‰spinneræˆ–å·²ç»æ¶ˆå¤±ï¼Œç»§ç»­
        await asyncio.sleep(0.3)
    
    async def init_page_for_aaaa(self):
        """åˆå§‹åŒ–é¡µé¢ï¼Œå¯¼èˆªå¹¶é€‰æ‹©AAAAè®°å½•ç±»å‹ï¼ˆåªæ‰§è¡Œä¸€æ¬¡ï¼‰"""
        logger.info("æ­£åœ¨åˆå§‹åŒ–DNSæŸ¥è¯¢é¡µé¢...")
        
        # å¯¼èˆªåˆ°DNSæŸ¥è¯¢é¡µé¢ï¼ˆä½¿ç”¨domcontentloadedè€Œä¸æ˜¯networkidleï¼Œæ›´å¿«ï¼‰
        await self.page.goto(self.base_url, wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(3)
        
        # é€‰æ‹©AAAAè®°å½•ç±»å‹ï¼ˆä½¿ç”¨force=Trueå¼ºåˆ¶ç‚¹å‡»ï¼Œå¿½ç•¥é®æŒ¡ï¼‰
        logger.info("é€‰æ‹©AAAAè®°å½•ç±»å‹...")
        dropdown = self.page.locator('div').filter(has_text=re.compile(r'^A$')).nth(1)
        await dropdown.click(force=True)
        await asyncio.sleep(0.5)
        
        # é€‰æ‹©AAAAé€‰é¡¹
        await self.page.get_by_title('AAAA').click(force=True)
        await asyncio.sleep(0.5)
        
        self.page_initialized = True
        logger.info("é¡µé¢åˆå§‹åŒ–å®Œæˆï¼Œå·²é€‰æ‹©AAAAè®°å½•ç±»å‹")
    
    async def check_if_blocked(self) -> bool:
        """æ£€æŸ¥æ˜¯å¦è¢«å°ç¦ï¼ˆ24å°æ—¶é™åˆ¶ï¼‰"""
        try:
            content = await self.page.content()
            if "è¯·æ±‚æ¬¡æ•°è¶…è¿‡é™åˆ¶" in content or "24å°æ—¶åé‡è¯•" in content:
                return True
        except:
            pass
        return False
    
    async def query_ipv6(self, domain: str, max_retries: int = 3) -> list:
        """æŸ¥è¯¢å•ä¸ªåŸŸåçš„IPv6åœ°å€ï¼ˆä¼˜åŒ–ç‰ˆï¼šç›´æ¥ä¿®æ”¹è¾“å…¥æ¡†ï¼Œä¸é‡æ–°åŠ è½½é¡µé¢ï¼‰"""
        ipv6_list = []
        
        for attempt in range(max_retries):
            try:
                # å¦‚æœé¡µé¢æœªåˆå§‹åŒ–ï¼Œå…ˆåˆå§‹åŒ–
                if not self.page_initialized:
                    await self.init_page_for_aaaa()
                
                # æ£€æŸ¥æ˜¯å¦è¢«å°ï¼ˆ24å°æ—¶é™åˆ¶ï¼‰
                if await self.check_if_blocked():
                    logger.warning("æ£€æµ‹åˆ°IPè¢«å°ç¦ï¼Œå°è¯•åˆ‡æ¢ä»£ç†èŠ‚ç‚¹...")
                    
                    # å°è¯•åˆ‡æ¢èŠ‚ç‚¹
                    next_proxy = self.get_next_proxy()
                    if next_proxy:
                        self.switch_clash_proxy(next_proxy)
                        await asyncio.sleep(3)
                        
                        # é‡æ–°åŠ è½½é¡µé¢
                        self.page_initialized = False
                        await self.init_page_for_aaaa()
                        
                        # å†æ¬¡æ£€æŸ¥æ˜¯å¦è¢«å°
                        if await self.check_if_blocked():
                            logger.error("=" * 60)
                            logger.error("åˆ‡æ¢èŠ‚ç‚¹åä»è¢«å°ç¦ï¼Œåœæ­¢çˆ¬è™«ï¼")
                            logger.error("è¯·æ‰‹åŠ¨æ›´æ¢ä»£ç†èŠ‚ç‚¹æˆ–ç­‰å¾…24å°æ—¶åå†è¯•")
                            logger.error("=" * 60)
                            raise Exception("IP_BLOCKED_24H")
                        else:
                            logger.info("åˆ‡æ¢èŠ‚ç‚¹æˆåŠŸï¼Œç»§ç»­çˆ¬å–...")
                    else:
                        logger.error("=" * 60)
                        logger.error("æ£€æµ‹åˆ°IPè¢«å°ç¦ï¼ˆ24å°æ—¶é™åˆ¶ï¼‰ï¼Œä¸”æ— æ³•åˆ‡æ¢èŠ‚ç‚¹ï¼Œåœæ­¢çˆ¬è™«ï¼")
                        logger.error("è¯·æ‰‹åŠ¨æ›´æ¢ä»£ç†èŠ‚ç‚¹æˆ–ç­‰å¾…24å°æ—¶åå†è¯•")
                        logger.error("=" * 60)
                        raise Exception("IP_BLOCKED_24H")
                
                logger.debug(f"å¼€å§‹æŸ¥è¯¢åŸŸå: {domain}")
                
                # ç›´æ¥ä¿®æ”¹è¾“å…¥æ¡†å†…å®¹ï¼ˆä½¿ç”¨force=Trueå¼ºåˆ¶æ“ä½œï¼‰
                input_box = self.page.get_by_role('textbox', name='ä¾‹ï¼šcn.bing.com')
                await input_box.click(force=True)
                await input_box.fill('')  # å…ˆæ¸…ç©º
                await input_box.fill(domain)
                await asyncio.sleep(0.3)
                
                # ç‚¹å‡»å¼€å§‹æµ‹è¯•
                await self.page.get_by_role('button', name='å¼€å§‹æµ‹è¯•').click(force=True)
                logger.debug(f"å·²ç‚¹å‡»å¼€å§‹æµ‹è¯•æŒ‰é’®")
                
                # ç­‰å¾…Loadingé®ç½©æ¶ˆå¤±ï¼ˆç­‰å¾…DNSæŸ¥è¯¢å®Œæˆï¼‰
                # å…³é”®ï¼šå¿…é¡»ç­‰å¾…è¿›åº¦æ¡åˆ°100%æˆ–Loadingæ¶ˆå¤±
                start_time = time.time()
                max_wait = 120  # æœ€å¤šç­‰å¾…120ç§’
                last_ipv6_count = 0
                stable_count = 0
                
                while time.time() - start_time < max_wait:
                    await asyncio.sleep(3)  # æ¯3ç§’æ£€æŸ¥ä¸€æ¬¡
                    elapsed = time.time() - start_time
                    
                    try:
                        # è·å–é¡µé¢å†…å®¹
                        content = await self.page.content()
                        
                        # æ£€æŸ¥Loadingæ˜¯å¦è¿˜åœ¨
                        # 1. æ£€æŸ¥Loadingæ–‡å­—
                        # 2. æ£€æŸ¥è¿›åº¦ç™¾åˆ†æ¯”ï¼ˆå¦‚æœä¸æ˜¯100%è¯´æ˜è¿˜åœ¨åŠ è½½ï¼‰
                        has_loading_text = 'Loading' in content
                        has_spinner = 'ant-spin-spinning' in content
                        
                        # æå–è¿›åº¦ç™¾åˆ†æ¯”
                        import re
                        progress_match = re.search(r'(\d+)%', content)
                        progress = int(progress_match.group(1)) if progress_match else 100
                        
                        is_loading = has_loading_text or has_spinner or (progress < 100)
                        
                        if is_loading:
                            logger.info(f"[{domain}] ç­‰å¾… {elapsed:.0f}s, DNSæŸ¥è¯¢è¿›è¡Œä¸­... (è¿›åº¦: {progress}%)")
                            continue
                        
                        # Loadingå®Œæˆåï¼Œå†ç­‰å¾…2ç§’ç¡®ä¿æ•°æ®æ¸²æŸ“å®Œæˆ
                        await asyncio.sleep(2)
                        content = await self.page.content()
                        
                        # æå–IPv6åœ°å€
                        valid_ipv6 = self.extract_ipv6_addresses(content)
                        current_count = len(valid_ipv6)
                        
                        logger.debug(f"[{domain}] Loadingå®Œæˆï¼Œæ‰¾åˆ° {current_count} ä¸ªIPv6")
                        
                        if current_count > 0:
                            if current_count == last_ipv6_count:
                                stable_count += 1
                                # å¦‚æœè¿ç»­2æ¬¡æ£€æŸ¥ç»“æœç¨³å®šï¼Œè®¤ä¸ºåŠ è½½å®Œæˆ
                                if stable_count >= 2:
                                    ipv6_list = list(valid_ipv6)
                                    logger.info(f"âœ“ [{domain}] æ‰¾åˆ° {len(ipv6_list)} ä¸ªIPv6: {', '.join(ipv6_list[:3])}{'...' if len(ipv6_list) > 3 else ''}")
                                    break
                            else:
                                stable_count = 0
                                last_ipv6_count = current_count
                        else:
                            # æ²¡æœ‰IPv6ï¼Œæ£€æŸ¥æ˜¯å¦æ˜¾ç¤ºæ— è®°å½•
                            if '0 ä¸ª IP' in content or '0ä¸ªIP' in content:
                                logger.info(f"- [{domain}] æ— IPv6è®°å½•")
                                break
                            # å¯èƒ½è¿˜åœ¨æ¸²æŸ“ï¼Œç»§ç»­ç­‰å¾…
                            stable_count += 1
                            if stable_count >= 3:
                                logger.info(f"- [{domain}] æ— IPv6è®°å½•ï¼ˆè¶…æ—¶ï¼‰")
                                break
                            
                    except Exception as e:
                        logger.warning(f"[{domain}] æ£€æŸ¥ç»“æœæ—¶å‡ºé”™: {e}")
                
                # å¦‚æœå¾ªç¯ç»“æŸä½†æœ‰ç»“æœï¼Œä¹Ÿè¿”å›
                if not ipv6_list and last_ipv6_count > 0:
                    content = await self.page.content()
                    valid_ipv6 = self.extract_ipv6_addresses(content)
                    ipv6_list = list(valid_ipv6)
                    if ipv6_list:
                        logger.info(f"âœ“ [{domain}] æ‰¾åˆ° {len(ipv6_list)} ä¸ªIPv6åœ°å€")
                
                # è¶…æ—¶ä½†æ²¡æœ‰ç»“æœ
                if not ipv6_list and time.time() - start_time >= max_wait:
                    logger.warning(f"âš  [{domain}] æŸ¥è¯¢è¶…æ—¶ï¼ˆ{max_wait}ç§’ï¼‰")
                        
                break  # æˆåŠŸå®Œæˆï¼Œé€€å‡ºé‡è¯•å¾ªç¯
                
            except Exception as e:
                logger.error(f"[{domain}] å°è¯• {attempt + 1}/{max_retries} å¤±è´¥: {e}")
                if attempt < max_retries - 1:
                    # é‡ç½®é¡µé¢çŠ¶æ€ï¼Œä¸‹æ¬¡é‡æ–°åˆå§‹åŒ–
                    self.page_initialized = False
                    await asyncio.sleep(3)
                    
        return ipv6_list
    
    def write_results_to_excel(self):
        """å°†ç»“æœå†™å…¥Excelæ–‡ä»¶"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        
        # æ‰¾åˆ°æˆ–åˆ›å»ºIPv6ç»“æœåˆ—
        # åŸæœ‰çš„AAAAè®°å½•åœ¨ç¬¬4åˆ—ï¼Œæˆ‘ä»¬åœ¨åé¢æ·»åŠ æ–°åˆ—
        header_row = list(ws[1])
        
        # æ‰¾åˆ°æœ€å¤§çš„IPv6æ•°é‡
        max_ipv6_count = max(len(ips) for ips in self.results.values()) if self.results else 0
        
        # æ·»åŠ æ–°çš„åˆ—æ ‡é¢˜ï¼ˆå¦‚æœéœ€è¦ï¼‰
        # ä»ç¬¬15åˆ—å¼€å§‹æ·»åŠ ï¼ˆå‡è®¾åŸæœ‰14åˆ—ï¼‰
        start_col = 15  # Oåˆ—å¼€å§‹
        
        # æ·»åŠ åˆ—æ ‡é¢˜
        for i in range(max_ipv6_count):
            col_letter = get_column_letter(start_col + i)
            ws[f'{col_letter}1'] = f'å®æ—¶IPv6_{i+1}'
        
        # å†™å…¥æ•°æ®
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1), start=2):
            domain = row[0].value
            if domain and domain.strip() in self.results:
                ipv6_list = self.results[domain.strip()]
                for i, ipv6 in enumerate(ipv6_list):
                    col_letter = get_column_letter(start_col + i)
                    ws[f'{col_letter}{row_idx}'] = ipv6
        
        # ä¿å­˜æ–‡ä»¶
        output_path = self.excel_path.replace('.xlsx', '_with_ipv6.xlsx')
        wb.save(output_path)
        wb.close()
        logger.info(f"ç»“æœå·²ä¿å­˜åˆ°: {output_path}")
        return output_path
    
    async def run(self, start_index: int = 0, end_index: int = None, resume: bool = True):
        """è¿è¡Œçˆ¬è™«
        
        Args:
            start_index: èµ·å§‹ç´¢å¼•
            end_index: ç»“æŸç´¢å¼•
            resume: æ˜¯å¦ä»ä¸Šæ¬¡ä¸­æ–­å¤„ç»§ç»­
        """
        logger.info("=" * 60)
        logger.info("DNS IPv6 è§£æçˆ¬è™« - å¼€å§‹è¿è¡Œ")
        logger.info("=" * 60)
        
        # è¯»å–åŸŸå
        domains = self.read_domains_from_excel()
        total = len(domains)
        logger.info(f"å…±è¯»å–åˆ° {total} ä¸ªåŸŸå")
        
        # å°è¯•åŠ è½½ä¹‹å‰çš„è¿›åº¦
        if resume:
            saved_index, saved_results = self.load_progress()
            if saved_index > start_index:
                start_index = saved_index
                self.results = saved_results
                logger.info(f"ä»ä¸Šæ¬¡ä¸­æ–­å¤„ç»§ç»­ï¼Œèµ·å§‹ç´¢å¼•: {start_index}")
        
        # å¤„ç†èŒƒå›´
        if end_index is None:
            end_index = total
        domains_to_process = domains[start_index:end_index]
        logger.info(f"å°†å¤„ç†ç¬¬ {start_index + 1} åˆ°ç¬¬ {end_index} ä¸ªåŸŸåï¼ˆå…± {len(domains_to_process)} ä¸ªï¼‰")
        
        # åˆå§‹åŒ–æµè§ˆå™¨
        await self.init_browser()
        
        success_count = 0
        fail_count = 0
        no_record_count = 0
        current_idx = start_index
        
        try:
            for idx, domain in enumerate(domains_to_process, start=start_index + 1):
                current_idx = idx
                
                # æ£€æŸ¥æ˜¯å¦éœ€è¦åˆ‡æ¢IP
                if self.use_proxy and self.current_request_count >= self.requests_per_ip:
                    await self.restart_browser_for_new_ip()
                
                logger.info(f"[{idx}/{end_index}] æ­£åœ¨æŸ¥è¯¢: {domain} (å½“å‰IPè¯·æ±‚æ•°: {self.current_request_count + 1}/{self.requests_per_ip})")
                query_start = time.time()
                
                ipv6_list = await self.query_ipv6(domain)
                self.results[domain] = ipv6_list
                self.current_request_count += 1
                
                query_time = time.time() - query_start
                
                if ipv6_list:
                    success_count += 1
                else:
                    no_record_count += 1
                
                logger.debug(f"[{domain}] æŸ¥è¯¢è€—æ—¶: {query_time:.1f}ç§’")
                
                # æ¯å¤„ç†10ä¸ªåŸŸåä¿å­˜ä¸€æ¬¡ä¸­é—´ç»“æœ
                if idx % 10 == 0:
                    logger.info(f"--- è¿›åº¦: {idx}/{end_index} ({idx*100//end_index}%) | æˆåŠŸ: {success_count} | æ— è®°å½•: {no_record_count} ---")
                    self.write_results_to_excel()
                
                # æ·»åŠ è¯·æ±‚é—´éš”ï¼ˆé¿å…è¢«å°ï¼‰
                await asyncio.sleep(3)  # æ¯ä¸ªè¯·æ±‚é—´éš”3ç§’
                
        except KeyboardInterrupt:
            logger.warning("ç”¨æˆ·ä¸­æ–­ï¼Œæ­£åœ¨ä¿å­˜è¿›åº¦å’Œç»“æœ...")
            self.save_progress(current_idx, self.results)
        except Exception as e:
            if "IP_BLOCKED_24H" in str(e):
                logger.error("å› IPè¢«å°ç¦ï¼ˆ24å°æ—¶é™åˆ¶ï¼‰è€Œåœæ­¢ï¼Œæ­£åœ¨ä¿å­˜è¿›åº¦...")
                self.save_progress(current_idx, self.results)
            else:
                logger.error(f"è¿è¡Œå‡ºé”™: {e}")
                fail_count += 1
                self.save_progress(current_idx, self.results)
        finally:
            # ä¿å­˜æœ€ç»ˆç»“æœ
            output_path = self.write_results_to_excel()
            
            # å…³é—­æµè§ˆå™¨
            await self.close_browser()
            
        logger.info("=" * 60)
        logger.info("çˆ¬è™«è¿è¡Œå®Œæˆ!")
        logger.info(f"ç»Ÿè®¡: æˆåŠŸè·å–IPv6: {success_count} | æ— IPv6è®°å½•: {no_record_count} | å¤±è´¥: {fail_count}")
        logger.info(f"ç»“æœæ–‡ä»¶: {output_path}")
        logger.info("=" * 60)
        
        return output_path


async def main():
    excel_path = "/Users/rongjiale/workspace/all_fobrain/new-project/åŸŸåèµ„äº§æ•°æ®_2025-12-23.xlsx"
    
    # åˆ›å»ºçˆ¬è™«å®ä¾‹
    # headless=False å¯ä»¥çœ‹åˆ°æµè§ˆå™¨æ“ä½œè¿‡ç¨‹ï¼Œè°ƒè¯•æ—¶ä½¿ç”¨
    # headless=True åå°è¿è¡Œï¼Œæ­£å¼ä½¿ç”¨
    # use_proxy=True ä½¿ç”¨æœ¬åœ°ä»£ç†
    # requests_per_ip=10 æ¯10ä¸ªè¯·æ±‚åˆ‡æ¢ä¸€æ¬¡IP
    crawler = DNSIPv6Crawler(
        excel_path, 
        headless=False,  # è®¾ç½®ä¸ºTrueå¯åå°è¿è¡Œ
        use_proxy=True,  # ä½¿ç”¨ä»£ç†
        proxy_host="127.0.0.1",
        proxy_port=7890,
        requests_per_ip=20  # æ¯20ä¸ªè¯·æ±‚åˆ‡æ¢IP
    )
    
    # è¿è¡Œçˆ¬è™«
    # resume=True ä¼šè‡ªåŠ¨ä»ä¸Šæ¬¡ä¸­æ–­å¤„ç»§ç»­
    # å¦‚æœæƒ³ä»å¤´å¼€å§‹ï¼Œè®¾ç½® resume=False
    await crawler.run(start_index=0, end_index=None, resume=True)


if __name__ == "__main__":
    asyncio.run(main())
